import types
import pandas as pd
import numpy as np
import datetime
import operator
import os
import os.path
import errno
import csv
import time
import datetime
import re
from openpyxl import load_workbook
from math import radians, cos, sin, asin, sqrt, fabs
# '02597a5b9e75cf6c' 0
# '02596c2b9e85bf69' 1
# '023d6dfce95b3997' 4
# '024a09d94ab01756' ------2------  025C411F85BAD2AC
# '024a0c2d305eb3cc' 6
# '0249cc55455bc6ad' 3
# '01ef64e4dec54025' 5

# in GPS file bus 2 and 7 are missing that cause empty files
VIRTUAL = 'virtualSchedule.csv'
DRIVER = 'driver11-17.csv'

SHUTTLE_NUM = ["763","747","762","748","787","788","528","526","[Deleted: ID 389]"]
SHUTTLE_INDEX = [0,1,2,3,4,5,6]
PHONE_ID = ["02597a5b9e75cf6c", "02596c2b9e85bf69", "024a09d94ab01756","0249cc55455bc6ad",
            "023d6dfce95b3997","01ef64e4dec54025","024a0c2d305eb3cc","025391f6de955621"]
NUM_OF_SHUTTLE = 8
STEP = 1
root = '/Users/fang/Desktop/sep/output/11_17/'

#PATH = r'/Users/fang/Desktop/sep/11_17'

UULat = 42.08687988
UULon = -75.96715875

UUendLat = 42.08687988
UUendLon = -75.96715875

GPS_dic = {}
with open('Gps.csv','rU') as csvfile:
	csvreader = csv.reader(csvfile)
	print 'reading GPS.csv...'
	for row in csvreader:
		str1 = ''
		if '\xef\xbb\xbf' in row[0]:
			str1 = row[0].replace('\xef\xbb\xbf','')
			row[0] = str1
			GPS_dic.setdefault(row[1], []).append(row)
		else:
			GPS_dic.setdefault(row[1], []).append(row)
print 'read GPS.csv end'

# Acc_dic = {}
# with open('Accelerometer.csv','rU') as csvfile:
# 	csvreader1 = csv.reader(csvfile)
# 	print 'reading Accelerometer.csv...'
# 	for row in csvreader1:
# 		str1 = ''
# 		if '\xef\xbb\xbf' in row[0]:
# 			str1 = row[0].replace('\xef\xbb\xbf','')
# 			row[0] = str1
# 			Acc_dic.setdefault(row[1], []).append(row)
# 		else:
# 			Acc_dic.setdefault(row[1], []).append(row)
# print 'read Accelerometer.csv end'

# Gry_dic = {}
# with open('Gyroscope.csv','rU') as csvfile:
# 	csvreader2 = csv.reader(csvfile)
# 	print 'reading Gyroscope.csv...'
# 	for row in csvreader2:
# 		str1 = ''
# 		if '\xef\xbb\xbf' in row[0]:
# 			str1 = row[0].replace('\xef\xbb\xbf','')
# 			row[0] = str1
# 			Gry_dic.setdefault(row[1], []).append(row)
# 		else:
# 			Gry_dic.setdefault(row[1], []).append(row)
# print 'read Gyroscope.csv end'

# Mag_dic = {}
# with open('Magnetometer.csv','rU') as csvfile:
# 	csvreader3 = csv.reader(csvfile)
# 	print 'reading Magnetometer.csv...'
# 	for row in csvreader3:
# 		str1 = ''
# 		if '\xef\xbb\xbf' in row[0]:
# 			str1 = row[0].replace('\xef\xbb\xbf','')
# 			row[0] = str1
# 			Mag_dic.setdefault(row[1], []).append(row)
# 		else:
# 			Mag_dic.setdefault(row[1], []).append(row)
# print 'read Magnetometer.csv end'

# Mot_dic = {}
# with open('MotionState.csv','rU') as csvfile:
# 	csvreader4 = csv.reader(csvfile)
# 	print 'reading MotionState.csv...'
# 	for row in csvreader4:
# 		str1 = ''
# 		if '\xef\xbb\xbf' in row[0]:
# 			str1 = row[0].replace('\xef\xbb\xbf','')
# 			row[0] = str1
# 			Mot_dic.setdefault(row[1], []).append(row)
# 		else:
# 			Mot_dic.setdefault(row[1], []).append(row)
# print 'read MotionState.csv end'

def get_result(sorted_list,route_dic):
    new_dic = {}
    exist_dic = {}
    # new_dic: dictionary that has everything
    # exist_dic : dictionary that only has bus[0,1,2,3,4,5,6,7]
    
    for number in range(len(sorted_list)):  #len(sorted_list)
        name = sorted_list[number][0]
        driver_slots = sorted_list[number][1]#NO.ith hardest driver schedule
        new_dic[name] = []
        exist_dic[name] = []
        #sort driver_slots according to time
        sorted_driver_slots = sorted(driver_slots, key=lambda item: date_to_timestamp(item[3] +' ' +convert_ap_pm(item[3], item[1].split('-')[0])))
        #=====================
        for slot in sorted_driver_slots:
        # ['Sunrise Pre-trip ALPHA', '6:30 AM - 7:00 AM', 'Jorge Fernando Flores', '2017/9/18']
            date = slot[3]
            time_split= slot[1].split('-')
            route1 = slot[0].upper()
            if route1.find('PARTIAL'):
                route = route1.split(' PARTIAL')[0]
            print '==========='
            print slot
            print route1
            print route
            print '==========='
            start = date + ' ' + convert_ap_pm(date, time_split[0])
            end = date + ' ' + convert_ap_pm(date, time_split[1])
            timestamp_start = date_to_timestamp(start)
            timestamp_end = date_to_timestamp(end)
            flag1 = 0
            for key in route_dic.keys():
                #if route.find(key) != -1 and key != 'L': #route match [['2017/9/5 7:23','763'], ['2017/9/5 7:23','763'] ,['2017/9/5 7:23','763']]
                if route.find(key) != -1 and key != 'L' and len(key) + route.index(key) == len(route): # dealwith CS & CS A problem
                    #print key
                    print '-------'
                    print route
                    print key
                    print '-------'
                    flag1 =1
                    candidate = []
                    breakpoint = []
                    for items in route_dic[key]:
                        if date_to_timestamp(items[0]) < timestamp_start:
                            candidate.append(items)
                        # if timestamp_start < date_to_timestamp(items[0]) < timestamp_end:
                        #     breakpoint.append(items)

                    if len(candidate) > 0:
                        flag = date_to_timestamp(candidate[0][0])
                        flag_index = 0
                        for i in range(len(candidate)):#from all smaller ones find the biggest one index as flag_index.
                            if date_to_timestamp(candidate[i][0]) > flag:
                                flag = date_to_timestamp(candidate[i][0])
                                flag_index = i
                        if len(breakpoint) == 0:
                            if len(candidate[flag_index][1]) < 5:
                                bus_index = SHUTTLE_NUM.index(candidate[flag_index][1])
                            else:
                                bus_index = 8
                            slot.append(bus_index)
                            slot.append(key)
                            break
                        else:#assume there is only one breakpoint
                            #print '!!!!!!!!!there is a break point!!!!!!!!!!!!!!'
                            slot[1] =  slot[1] + breakpoint[0][0]

                            if len(candidate[flag_index][1]) < 5:
                                bus_index = SHUTTLE_NUM.index(candidate[flag_index][1])
                            else:
                                bus_index = 8
                            slot.append(bus_index)

                            if len(breakpoint[0][1]) < 5:
                                break_index = SHUTTLE_NUM.index(breakpoint[0][1])
                            else:
                                break_index = 8
                            slot.append(break_index)
                            slot.append(key)
                            break
                    else:
                        slot.append('no bus found')
                        break
            if flag1 == 0:
                slot.append("route not found")
            new_dic[name].append(slot)
            if slot[-2] in SHUTTLE_INDEX:
                exist_dic[name].append(slot)

    return new_dic, exist_dic

def hav(theta):
	s = sin(theta / 2)
	return s * s
def get_distance(lat0, lng0, lat1, lng1):
	r = 6371
	lat0 = radians(lat0)
	lat1 = radians(lat1)
	lng0 = radians(lng0)
	lng1 = radians(lng1)

	dlng = fabs(lng0 - lng1)
	dlat = fabs(lat0 - lat1)
	h = hav(dlat) + cos(lat0) * cos(lat1) * hav(dlng)
	distance = 2 * r * asin(sqrt(h)) * 1000
	return distance     # return meter

def date_to_timestamp(date):  # convert Y/M/D H:M to x xxx xxx xxx . xxx
    time_array = time.strptime(date, "%Y/%m/%d %H:%M")
    timestamp = time.mktime(time_array)
    return timestamp

def timestamp_to_date(timestamp):  # convert x xxx xxx xxx . xxx to Y/M/D H:M
    time_local = time.localtime(timestamp)
    dt = time.strftime("%Y/%-m/%-d %-H:%-M", time_local)  # %-m %-d if it's 07/06, will return 7/6
    return dt

def convert_ap_pm(date, time):
    """
    :param date: eg: 2017/7/7
    :param time: eg: 1pm
    :return:     eg: 13:00
    """
    temp = date + ' ' + time
    # print temp
    res = pd.to_datetime(temp).strftime('%-H:%-M')
    return res
def get_other (startime,endtime,busnumber,filename,wfilename,kind,index,dictname):
	"""
	8663665	024a09d94ab01756	1505882567526.00	0.337697774	0.083825685	9.649534225
	"""
	if index == 1:
		temp_dic = Acc_dic
	if index == 2:
		temp_dic = Gry_dic
	if index == 3:
		temp_dic = Mag_dic
	if index == 4:
		temp_dic = Mot_dic

	flag1 = 0
	print 'geting ' + kind +'info......'
	print '===infunction==='
	print timestamp_to_date(startime)
	print timestamp_to_date(endtime)
	print '===infunction_end==='
	path = root+dictname
	filepath = root+dictname+'/'+wfilename + kind + ".csv"
	if not os.path.exists(os.path.dirname(filepath)):
		try:
			os.makedirs(os.path.dirname(filepath))
		except OSError as exc:
			if exc.errno != errno.EEXIST:
				raise
				
	with open(filepath,"w") as csvfile:
		ww = csv.writer(csvfile)
		if temp_dic.has_key(PHONE_ID[busnumber]):
			buslist = temp_dic[PHONE_ID[bus]]
			for i in range(0,len(buslist),STEP):
				y = int(buslist[i][2])/1000
				# x = timestamp_to_date(y)
				if y >= startime and y <= endtime:
					row = [(buslist[i])]
					ww.writerows(row)
					flag1 = 1
			if flag1 == 0:
				ww.writerows([('no data during this time slot! in '+ filename)])
				print 'no data during this time slot! in '+ filename
		else:
			whichbus = str(busnumber)
			ww.writerows([('In Sep bus '+ whichbus +' has no data')])
			print 'This bus has no data in' + filename + '!'

def Wfile(schedule,bus,date,dictname):

	for slot in schedule:
		subroute = slot[2].replace('/','')
		wfilename = date.replace('/','.') + '-' + slot[0] + subroute

		start = date + ' ' + slot[0]
		end = date + ' ' + slot[1]
		timestamp_start = date_to_timestamp(start)
		timestamp_end = date_to_timestamp(end)
		temp = []
		flag = 0
		# get timestamp_start and timestamp_end
		if GPS_dic.has_key(PHONE_ID[bus]):
			buslist = GPS_dic[PHONE_ID[bus]]
			starflag = 0 #locker
			endflag = 0
			newstartStamp = timestamp_start
			newendStamp = timestamp_end
			print timestamp_to_date(timestamp_start)
			print timestamp_to_date(timestamp_end)
			for i in range(0,len(buslist),STEP):
				y = int(buslist[i][2])/1000
				if y >= timestamp_start and y < timestamp_end:
					#lastone = buslist[i]
					startdistance = get_distance(UULat,UULon,float(buslist[i][3]),float(buslist[i][4]))
					if starflag == 0:
						if startdistance > 350 :
							timestamp_start = timestamp_start + 60
							print 'add 60s'
					if startdistance < 300 and starflag == 0:
						starflag = 1
						newstartStamp = y
						print 'reset-start'
						print timestamp_to_date(newstartStamp)

				elif y >= timestamp_end and endflag == 0:
					new = get_distance(UUendLat,UUendLon,float(buslist[i][3]),float(buslist[i][4]))
					if new > 350:
						timestamp_end = timestamp_end + 60
					if new < 300:
						newendStamp = y
						print 'reset-end'
						print timestamp_to_date(newendStamp)
						endflag = 1
		path = root+dictname
		filepath = root+dictname+'/'+wfilename+"_GPS.csv"
		filepath1 = root+dictname+'/'+wfilename+"_Web.csv"
		if not os.path.exists(os.path.dirname(filepath)):
			try:
				os.makedirs(os.path.dirname(filepath))
			except OSError as exc:
				if exc.errno != errno.EEXIST:
					raise
		if not os.path.exists(os.path.dirname(filepath1)):
			try:
				os.makedirs(os.path.dirname(filepath1))
			except OSError as exc:
				if exc.errno != errno.EEXIST:
					raise
		with open( filepath , "w") as csvfile1, open(filepath1, "w") as web:
		#with open(filepath1, "w") as web:
			writergps = csv.writer(csvfile1)
			webR = csv.writer(web)
			webR.writerow(["lat","lng","comment"])

			if GPS_dic.has_key(PHONE_ID[bus]):
				buslist = GPS_dic[PHONE_ID[bus]]
				print 'newnewnewnewnew'
				print timestamp_to_date(newstartStamp)
				print timestamp_to_date(newendStamp)
				for i in range(0,len(buslist),STEP):
					y = int(buslist[i][2])/1000
					x = timestamp_to_date(y)
					if y >= newstartStamp and y <= newendStamp:
						temp.append(buslist[i])
						row = [(buslist[i])]
						row_web = [(float(buslist[i][3]),float(buslist[i][4]))]
						writergps.writerows(row)
						webR.writerows(row_web)
						flag = 1

				if flag == 0:
					writergps.writerows([('no data during this time slot!')])
					webR.writerows([('no data during this time slot!')])
					print 'no data during this time slot!'
				if flag == 1:
					print 'result'
					print timestamp_to_date(int(temp[0][2])/1000)
					print timestamp_to_date(int(temp[-1][2])/1000)
					print get_distance(UULat,UULon,float(temp[0][3]),float(temp[0][4]))
					print get_distance(UUendLat,UUendLon,float(temp[-1][3]),float(temp[-1][4]))
				print 'data write to '+wfilename+".csv"+'\n'
			else:
				string = str(bus)
				writergps.writerows([('in Sep BUS '+string+' do not have data')])
				webR.writerows([('in Sep BUS '+string+' do not have data')])
				print 'This bus has no data in GPS file!'
		# get_other(newstartStamp,newendStamp,bus,'Accelerometer.csv',wfilename,'_Acc',1,dictname)
		# get_other(newstartStamp,newendStamp,bus,'Gyroscope.csv',wfilename,'_Gyr',2,dictname)
		# get_other(newstartStamp,newendStamp,bus,'Magnetometer.csv',wfilename,'_Mag',3,dictname)
		# get_other(newstartStamp,newendStamp,bus,'MotionState.csv',wfilename,'_Mot',4,dictname)
		
def getDictxlsx(day,Stime):
	wb = load_workbook("Fall.xlsx")
	# print(wb.sheetnames)
	if day>=1 and day <=4:
		sheetname = "MON-THURSDAY"
	if day == 5 and Stime >= 1200:
		sheetname = "FRIDAY NIGHTS"
	if day == 5 and Stime < 1200:
		sheetname = "MON-THURSDAY"
	if day == 6:
		sheetname = "SATURDAY"
	if day == 7:
		sheetname = "SUNDAY"
	sheet = wb.get_sheet_by_name(sheetname)
	dic = {}
	'''
	dic  = {ALPHA:[6:30,6:55,PRETRIP] }
	'''
	temp = sorted(sheet.merged_cell_ranges, key=lambda item: item[0], reverse=False)
	for y in temp:
		merge = y.split(':')
		# merge[0] = 'B2' merge[1] = 'B7'
		route = sheet[merge[0][0]+'1'].value.encode()
		Ltime_temp = sheet['A'+merge[0][1:]].value
		Rtime_temp = sheet['A'+merge[1][1:]].value
		old = str(sheet[merge[0]].value)
		subroute = re.sub(r"\s{2,}", " ", old)

		if isinstance(Ltime_temp,datetime.time) : #type 'datetime.time'
			Lctime = str(Ltime_temp)[0:5]
		elif isinstance(Ltime_temp,datetime.datetime): #type 'datetime.datetime'
			Lctime = str(Ltime_temp)[11:16]
		if isinstance(Rtime_temp,datetime.time) : #type 'datetime.time'
			Rctime = str(Rtime_temp)[0:5]
		elif isinstance(Rtime_temp,datetime.datetime): #type 'datetime.datetime'
			Rctime = str(Rtime_temp)[11:16]
		#print route,Lctime,Rctime,subroute
		dic.setdefault(route, []).append([Lctime,Rctime,subroute])
	return dic
def get_mint(hhmm):
	temp = hhmm.split(':')
	hour = int(temp[0])
	mint = int(temp[1])
	return hour*60 + mint

if __name__=="__main__":

	for key in GPS_dic.keys():
		print key

	virtialRouteTable = []
	route_dic = {}
	with open(VIRTUAL,'rU') as csvfile:
		csvreader1 = csv.reader(csvfile)
		for row in csvreader1:
			str1 = ''
			if '\xef\xbb\xbf' in row[0]:
				str1 = row[0].replace('\xef\xbb\xbf','')
				row[0] = str1
				virtialRouteTable.append(row)
			else:
				virtialRouteTable.append(row)
    #build dictionary
	for i in range(1,len(virtialRouteTable)):
		if virtialRouteTable[i][4].find("Details:") != -1:
			a = virtialRouteTable[i][4].split("Details: ")
			if len(a) > 1:
				temp = [virtialRouteTable[i][0],virtialRouteTable[i][2]]#add time and BUS number e.g['2017/9/5 7:23','763']
				x = a[1].upper()
				route_dic.setdefault(x, []).append(temp)

	virtualTable = []
	dic_driver = {}
	with open(DRIVER,'rU') as csvfile:
		csvreader = csv.reader(csvfile)
		for row in csvreader:
			str1 = ''
			if '\xef\xbb\xbf' in row[0]:
				str1 = row[0].replace('\xef\xbb\xbf','')
				row[0] = str1
				virtualTable.append(row)
			else:
				virtualTable.append(row)

	for i in range(1,len(virtualTable)):
    #['Jorge Fernando Flores', '2017/09/11', 'Monday', 'Sunrise Pre-trip ALPHA', '6:30 AM-7:00 AM', '0.5', 'details', '']
		if virtualTable[i][7] == '':
			temp =[]
			newtime = virtualTable[i][4].replace('-',' - ')
			temp.extend([virtualTable[i][3], newtime ,virtualTable[i][0],virtualTable[i][1]])
			x = temp[2].replace(' ','')
			if x.isalpha():
				dic_driver.setdefault(temp[2], []).append(temp)
		elif virtualTable[i][7].find('Unapproved') == -1:
			if virtualTable[i][7].find(':') != -1:
				new = virtualTable[i][7].split(': ')
				newDriver = new[1].replace('.','')
				temp =[]
				newtime = virtualTable[i][4].replace('-',' - ')
				temp.extend([virtualTable[i][3], newtime ,newDriver,virtualTable[i][1]])
				x = temp[2].replace(' ','')
				if x.isalpha():
					dic_driver.setdefault(temp[2], []).append(temp)

	sorted_list1 = sorted(dic_driver.items(), key=lambda item: len(item[1]), reverse=True)
	result,exist_result= get_result(sorted_list1,route_dic)
	sorted_result = sorted(exist_result.items(), key=lambda item: len(item[1]), reverse=True)

	"""
	['6704101', '024a09d94ab01756', '1505667614326', '42.10743427', '-75.98704116', '0', '0', '0']
	   0                 1                  2                3              4         
	 GPS file
	"""
	"""
	['Eve CHI', '10:00 PM - 12:30 AM', 'Administrator', '2017/09/12', 7] result from driverlist
	"""
	# source = [ ['Sunrise Pre-trip BETA', '6:30 AM - 7:00 AM', 'Reece Pyankaroo', '2017/09/11', 4, 'BETA'],
	# 			['Sunrise CHI', '7:00 AM - 9:00 AM', 'Victor Santana Aguilera', '2017/09/13', 5, 'CHI'] ]

	for i in range(len(sorted_result)):
		driverID = str(i) + str(sorted_result[i][0])
		source = sorted_result[i][1]
		for test in source:
			result = []
			bus = test[-2]
			date = test[3]
			times = test[1].split('-')
			start = convert_ap_pm(date, times[0])
			end = convert_ap_pm(date, times[1])
			route = test[-1]
			print start,end,route
			S = get_mint(start)
			E = get_mint(end)
			print '======'
			week = datetime.datetime.strptime(date,"%Y/%m/%d").weekday() + 1
			print week
			Excel = getDictxlsx(week,S)
			for x in Excel[route]:
				if  get_mint(x[0]) >= S and get_mint(x[0]) < E:
					print x
					result.append(x)
				if S > E:
					if get_mint(x[0]) >= S or get_mint(x[0]) < E:
						print x
						result.append(x)
			print '\n'
			Wfile(result,bus,date,driverID)

















	# for test in source:
		
	# 	result = []
	# 	bus = test[-2]
	# 	date = test[3]
	# 	times = test[1].split('-')
	# 	start = convert_ap_pm(date, times[0])
	# 	end = convert_ap_pm(date, times[1])
	# 	route = test[-1]
	# 	print start,end,route
	# 	S = get_mint(start)
	# 	E = get_mint(end)
	# 	print '======'
	# 	week = datetime.datetime.strptime(date,"%Y/%m/%d").weekday() + 1
	# 	print week
	# 	Excel = getDictxlsx(week,S)
	# 	for x in Excel[route]:
	# 		if  get_mint(x[0]) >= S and get_mint(x[0]) < E:
	# 			print x
	# 			result.append(x)
	# 		if S > E:
	# 			if get_mint(x[0]) >= S or get_mint(x[0]) < E:
	# 				print x
	# 				result.append(x)
	# 	print '\n'
	# 	Wfile(result,bus,date,driverID)





