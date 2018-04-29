import types
import pandas as pd
import numpy as np
import datetime
import operator
import os
import csv
import time
import datetime
import re
from openpyxl import load_workbook
from math import radians, cos, sin, asin, sqrt, fabs
# '02597a5b9e75cf6c' 0
# '02596c2b9e85bf69' 1
# '023d6dfce95b3997' 4
# '024a09d94ab01756' ------2
# '024a0c2d305eb3cc' 6
# '0249cc55455bc6ad' 3
# '01ef64e4dec54025' 5

# in GPS file bus 2 and 7 are missing that cause empty files

SHUTTLE_NUM = ["763","747","762","748","787","788","528","526","[Deleted: ID 389]"]
SHUTTLE_INDEX = [0,1,2,3,4,5,6,7]
PHONE_ID = ["02597a5b9e75cf6c", "02596c2b9e85bf69", "024a09d94ab01756","0249cc55455bc6ad",
            "023d6dfce95b3997","01ef64e4dec54025","024a0c2d305eb3cc","025391f6de955621"]
NUM_OF_SHUTTLE = 8

# UULat = 42.086942806500474
# UULon = -75.96704414865724

UULat = 42.08687988
UULon = -75.96715875

UUendLat = 42.08687988
UUendLon = -75.96715875

def hav(theta):
	s = sin(theta / 2)
	return s * s
def get_distance(lat0, lng0, lat1, lng1):
	r = 6371
	lat0 = radians(lat0)
	lat1 = radians(lat1)
	lng0 = radians(lng0)
	lng1 = radians(lng1)

	dlng = fabs(lng0 - lng1)
	dlat = fabs(lat0 - lat1)
	h = hav(dlat) + cos(lat0) * cos(lat1) * hav(dlng)
	distance = 2 * r * asin(sqrt(h)) * 1000
	return distance     # return meter

def date_to_timestamp(date):  # convert Y/M/D H:M to x xxx xxx xxx . xxx
    time_array = time.strptime(date, "%Y/%m/%d %H:%M")
    timestamp = time.mktime(time_array)
    return timestamp

def timestamp_to_date(timestamp):  # convert x xxx xxx xxx . xxx to Y/M/D H:M
    time_local = time.localtime(timestamp)
    dt = time.strftime("%Y/%-m/%-d %-H:%-M", time_local)  # %-m %-d if it's 07/06, will return 7/6
    return dt

def convert_ap_pm(date, time):
    """
    :param date: eg: 2017/7/7
    :param time: eg: 1pm
    :return:     eg: 13:00
    """
    temp = date + ' ' + time
    # print temp
    res = pd.to_datetime(temp).strftime('%-H:%-M')
    return res
def get_other (startime,endtime,busnumber,routename,filename,kind,day):
	"""
	8663665	024a09d94ab01756	1505882567526.00	0.337697774	0.083825685	9.649534225
	"""
	print 'geting ' + kind +'info......'
	print '===infunction==='
	print timestamp_to_date(startime)
	print timestamp_to_date(endtime)
	print '===infunction_end==='
	temp_dic = {}
	with open(filename,'rU') as csvfile:
	        csvreader1 = csv.reader(csvfile)
	        for row in csvreader1:
	            str1 = ''
	            if '\xef\xbb\xbf' in row[0]:
	                str1 = row[0].replace('\xef\xbb\xbf','')
	                row[0] = str1
	                temp_dic.setdefault(row[1], []).append(row)
	            else:
	            	temp_dic.setdefault(row[1], []).append(row)
	flag = 0
	with open(routename+ kind + day + ".csv","w") as csvfile:
		ww = csv.writer(csvfile)
		if temp_dic.has_key(PHONE_ID[busnumber]):
			buslist = temp_dic[PHONE_ID[bus]]
			for i in range(0,len(buslist),30):
				y = int(buslist[i][2])/1000
				# x = timestamp_to_date(y)
				if y >= startime and y <= endtime:
					row = [(buslist[i])]
					ww.writerows(row)
					flag = 1
			if falg == 0:
				print 'no data during this time slot! in '+ filename

		else:
			print 'This bus has no data in' + filename + '!'

def Wfile(dic,schedule,bus,date):
	for slot in schedule:
		subroute = slot[2].replace('/','')
		wfilename = date.replace('/','.') + '-' + slot[0] + subroute

		start = date + ' ' + slot[0]
		end = date + ' ' + slot[1]
		timestamp_start = date_to_timestamp(start)
		timestamp_end = date_to_timestamp(end)
		temp = []
		flag = 0
		# get timestamp_start and timestamp_end
		if dic.has_key(PHONE_ID[bus]):
			buslist = dic[PHONE_ID[bus]]
			starflag = 0 #locker
			endflag = 0
			newstartStamp = timestamp_start
			newendStamp = timestamp_end
			print timestamp_to_date(timestamp_start)
			print timestamp_to_date(timestamp_end)
			for i in range(0,len(buslist),30):
				y = int(buslist[i][2])/1000
				if y >= timestamp_start and y < timestamp_end:
					#lastone = buslist[i]
					startdistance = get_distance(UULat,UULon,float(buslist[i][3]),float(buslist[i][4]))
					if starflag == 0:
						if startdistance > 350 :
							timestamp_start = timestamp_start + 60
							print 'add 60s'
					if startdistance < 300 and starflag == 0:
						starflag = 1
						newstartStamp = y
						print 'reset-start'
						print timestamp_to_date(newstartStamp)

				elif y >= timestamp_end and endflag == 0:
					new = get_distance(UUendLat,UUendLon,float(buslist[i][3]),float(buslist[i][4]))
					if new > 350:
						timestamp_end = timestamp_end + 60
					if new < 300:
						newendStamp = y
						print 'reset-end'
						print timestamp_to_date(newendStamp)
						endflag = 1
		with open(wfilename+"_GPS.csv","w") as csvfile1, open(wfilename+"_Web.csv","w") as web:
			writer = csv.writer(csvfile1)
			webR = csv.writer(web)
			webR.writerow(["lat","lng","comment"])

			if dic.has_key(PHONE_ID[bus]):
				buslist = dic[PHONE_ID[bus]]
				print 'newnewnewnewnew'
				print timestamp_to_date(newstartStamp)
				print timestamp_to_date(newendStamp)
				for i in range(0,len(buslist),30):
					y = int(buslist[i][2])/1000
					x = timestamp_to_date(y)
					if y >= newstartStamp and y <= newendStamp:
						temp.append(buslist[i])
						row = [(buslist[i])]
						row_web = [(float(buslist[i][3]),float(buslist[i][4]))]
						writer.writerows(row)
						webR.writerows(row_web)
						flag = 1

				if flag == 0:
					print 'no data during this time slot!'
				if flag == 1:
					print 'result'
					print timestamp_to_date(int(temp[0][2])/1000)
					print timestamp_to_date(int(temp[-1][2])/1000)
					print get_distance(UULat,UULon,float(temp[0][3]),float(temp[0][4]))
					print get_distance(UUendLat,UUendLon,float(temp[-1][3]),float(temp[-1][4]))
				print 'data write to '+wfilename+".csv"+'\n'
			else:
				print 'This bus has no data in GPS file!'
		#get_other(newstartStamp,newendStamp,bus,route,filename,kind,date)
		
def getDict(day,Stime):
	wb = load_workbook("Fall.xlsx")
	# print(wb.sheetnames)
	if day>=1 and day <=4:
		sheetname = "MON-THURSDAY"
	if day == 5 and Stime >= 1200:
		sheetname = "FRIDAY NIGHTS"
	if day == 5 and Stime < 1200:
		sheetname = "MON-THURSDAY"
	if day == 6:
		sheetname = "SATURDAY"
	if day == 7:
		sheetname = "SUNDAY"

	sheet = wb.get_sheet_by_name(sheetname)
	dic = {}
	'''
	dic  = {ALPHA:[6:30,6:55,PRETRIP] }
	'''
	temp = sorted(sheet.merged_cell_ranges, key=lambda item: item[0], reverse=False)
	for y in temp:
		merge = y.split(':')
		# merge[0] = 'B2' merge[1] = 'B7'
		route = sheet[merge[0][0]+'1'].value.encode()
		Ltime_temp = sheet['A'+merge[0][1:]].value
		Rtime_temp = sheet['A'+merge[1][1:]].value
		old = str(sheet[merge[0]].value)
		subroute = re.sub(r"\s{2,}", " ", old)

		if isinstance(Ltime_temp,datetime.time) : #type 'datetime.time'
			Lctime = str(Ltime_temp)[0:5]
		elif isinstance(Ltime_temp,datetime.datetime): #type 'datetime.datetime'
			Lctime = str(Ltime_temp)[11:16]
		if isinstance(Rtime_temp,datetime.time) : #type 'datetime.time'
			Rctime = str(Rtime_temp)[0:5]
		elif isinstance(Rtime_temp,datetime.datetime): #type 'datetime.datetime'
			Rctime = str(Rtime_temp)[11:16]
		#print route,Lctime,Rctime,subroute
		dic.setdefault(route, []).append([Lctime,Rctime,subroute])
	return dic
def get_mint(hhmm):
	temp = hhmm.split(':')
	hour = int(temp[0])
	mint = int(temp[1])
	return hour*60 + mint

if __name__=="__main__":
	GPSdic = {}
	with open('Gps.csv','rU') as csvfile:
		csvreader = csv.reader(csvfile)
		for row in csvreader:
			str1 = ''
			if '\xef\xbb\xbf' in row[0]:
				str1 = row[0].replace('\xef\xbb\xbf','')
				row[0] = str1
				GPSdic.setdefault(row[1], []).append(row)
			else:
				GPSdic.setdefault(row[1], []).append(row)
	# for key in GPSdic.keys():
	# 	print key

	"""
	['6704101', '024a09d94ab01756', '1505667614326', '42.10743427', '-75.98704116', '0', '0', '0']
	   0                 1                  2                3              4         
	 GPS file
	"""
	"""
	['Eve CHI', '10:00 PM - 12:30 AM', 'Administrator', '2017/09/12', 7] result from driverlist
	"""
	# source = [['Sunset EPSILON', '5:00 PM - 7:15 PM', 'Administrator', '2017/09/11', 5, 'EPSILON'],
	# ['Mid GAMMA', '11:00 AM - 1:00 PM', 'Administrator', '2017/09/12', 5, 'GAMMA'],
	# ['Mid CS B', '11:15 AM - 1:30 PM', 'Administrator', '2017/09/12', 2, 'CS B'],
	# ['Eve CHI', '10:00 PM - 12:30 AM', 'Administrator', '2017/09/12', 7, 'CHI'],
	# ['Noon CS A', '1:00 PM - 3:00 PM', 'Administrator', '2017/09/13', 0, 'CS A'],
	# ['Sunset EPSILON', '5:00 PM - 7:15 PM', 'Administrator', '2017/09/13', 4, 'EPSILON'],
	# ['Mid CS B', '11:15 AM - 1:30 PM', 'Administrator', '2017/09/15', 1, 'CS B'],
	# ['Afternoon BETA', '2:00 PM - 4:15 PM', 'Administrator', '2017/09/15', 4, 'BETA'],
	# ['Sunset CS C', '6:00 PM - 8:00 PM', 'Administrator', '2017/09/15', 3, 'CS C'],
	# ['PM BETA', '6:45 PM - 9:30 PM', 'Administrator', '2017/09/15', 4, 'BETA'],
	# ['Sunset CS B partial 2', '7:15 PM - 7:30 PM', 'Administrator', '2017/09/15', 1, 'CS B'] ]

	source = [['Sunset EPSILON', '5:00 PM - 7:15 PM', 'Administrator', '2017/09/11', 5, 'EPSILON']]
	for test in source:
		result = []
		bus = test[-2]
		date = test[3]
		times = test[1].split('-')
		start = convert_ap_pm(date, times[0])
		end = convert_ap_pm(date, times[1])
		route = test[-1]
		print start,end,route
		S = get_mint(start)
		E = get_mint(end)
		print '======'
		week = datetime.datetime.strptime(date,"%Y/%m/%d").weekday() + 1
		print week
		Excel = getDict(week,S)
		for x in Excel[route]:
			if  get_mint(x[0]) >= S and get_mint(x[0]) < E:
				print x
				result.append(x)
			if S > E:
				if get_mint(x[0]) >= S or get_mint(x[0]) < E:
					print x
					result.append(x)
		print '\n'
		Wfile(GPSdic,result,bus,date)





